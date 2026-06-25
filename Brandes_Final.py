import tkinter as tk

import openpyxl as op

from tkinter import ttk, messagebox

from datetime import date

import os



if not os.path.exists("Brandes_Database.xlsx"):



    workbook = op.Workbook()

    sheet = workbook.active

    sheet.title = "Printing Orders"



    headings = ["Order ID","Customer Name","Print Type","Paper Size","Quantity","Price Per Page","Total Price","Status","Date Ordered"]



    sheet.append(headings)

    workbook.save("Brandes_Database.xlsx")



def display():



    workbook = op.load_workbook("Brandes_Database.xlsx")

    sheet = workbook.active



    if sheet.cell(row=1, column=9).value is None:

        sheet.cell(row=1, column=9).value = "Date Ordered"



        for r in range(2, sheet.max_row + 1):

            sheet.cell(row=r, column=9).value = "No Date"



        workbook.save("Brandes_Database.xlsx")



    for row in table.get_children():

        table.delete(row)



    for row in sheet.iter_rows(min_row=2, values_only=True):



        row = list(row)



        while len(row) < 9:

            row.append("No Date")



        table.insert("", tk.END, values=row)

def input_validation():



    cstmr = customer_entry.get()

    qnt = quantity_entry.get()

    prc = price_entry.get()



    if not cstmr or not qnt or not prc:

        messagebox.showerror("Error", "All fields are required!")

        return False



    if not qnt.isdigit():

        messagebox.showerror("Error", "Quantity must be a number!")

        return False



    try:

        float(prc)



    except:

        messagebox.showerror("Error", "Price must be a number!")

        return False

    

    if print_type_var.get() == "Select Print Type":

        messagebox.showerror("Error", "Please select a print type!")

        return False



    if paper_size_var.get() == "Select Paper Size":

        messagebox.showerror("Error", "Please select a paper size!")

        return False



    if status_var.get() == "Select Status":

        messagebox.showerror("Error", "Please select a status!")

        return False



    return True







def saving():



    if not input_validation():

        return



    cstomer = customer_entry.get()

    prnt_type = print_type_var.get()

    ppr_size = paper_size_var.get()

    quant = int(quantity_entry.get())

    price = float(price_entry.get())

    status = status_var .get()



    total = quant * price



    date_ordered = str(date.today())



    workbook = op.load_workbook("Brandes_Database.xlsx")

    sheet = workbook.active



    new_id = "ORD-" + str(sheet.max_row)



    sheet.append([

        new_id,

        cstomer,

        prnt_type,

        ppr_size,

        quant,

        price,

        total,

        status,

        date_ordered

    ])



    workbook.save("Brandes_Database.xlsx")



    messagebox.showinfo("Success", "Record added successfully!")



    clear_fields()

    display()





def auto_populate(event):



    selected = table.focus()

    values = table.item(selected, "values")



    if values:



        customer_entry.delete(0, tk.END)

        quantity_entry.delete(0, tk.END)

        price_entry.delete(0, tk.END)



        customer_entry.insert(0, values[1])

        print_type_var.set(values[2])

        paper_size_var.set(values[3])

        quantity_entry.insert(0, values[4])

        price_entry.insert(0, values[5])

        status_var.set(values[7])



def update():



    selected = table.focus()



    if not selected:

        messagebox.showerror("Error", "Select a record first!")

        return



    if not input_validation():

        return



    values = table.item(selected, "values")

    record_id = values[0]



    customer = customer_entry.get()

    print_type = print_type_var.get()

    paper_size = paper_size_var.get()

    quantity = int(quantity_entry.get())

    price = float(price_entry.get())

    status = status_var.get()



    total = quantity * price



    workbook = op.load_workbook("Brandes_Database.xlsx")

    sheet = workbook.active



    for rows in sheet.iter_rows(min_row=2):



        if str(rows[0].value) == str(record_id):



            rows[1].value = customer

            rows[2].value = print_type

            rows[3].value = paper_size

            rows[4].value = quantity

            rows[5].value = price

            rows[6].value = total

            rows[7].value = status



    workbook.save("Brandes_Database.xlsx")



    messagebox.showinfo("Success", "Record updated successfully!")



    clear_fields()

    display()





def delete():



    selected = table.focus()



    if not selected:

        messagebox.showerror("Error", "Select a record first!")

        return



    values = table.item(selected, "values")

    record_id = values[0]



    confirm = messagebox.askyesno(

        "Confirm",

        "Are you sure you want to delete this record?"

    )



    if not confirm:

        return



    workbook = op.load_workbook("Brandes_Database.xlsx")

    sheet = workbook.active



    for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):



        if str(row[0].value) == str(record_id):

            sheet.delete_rows(i)

            break



    workbook.save("Brandes_Database.xlsx")



    messagebox.showinfo("Success", "Record deleted successfully!")



    clear_fields()

    display()





def search_record():



    search = search_entry.get().lower()



    for row in table.get_children():

        table.delete(row)



    workbook = op.load_workbook("Brandes_Database.xlsx")

    sheet = workbook.active



    for row in sheet.iter_rows(min_row=2, values_only=True):



        if search in str(row).lower():

            table.insert("", tk.END, values=row)





def print_receipt():



    selected = table.focus()



    if not selected:

        messagebox.showerror("Error", "Select a record first!")

        return



    values = table.item(selected, "values")



    if len(values) > 8:

        date_ordered = values[8]

    else:

        date_ordered = "No Date"



    receipt = f"""



------------ PRINT RECEIPT ---------------



Order ID: {values[0]}

Customer Name: {values[1]}

Print Type: {values[2]}

Paper Size: {values[3]}

Quantity: {values[4]}

Price Per Page: ₱{values[5]}

Total Price: ₱{values[6]}

Status: {values[7]}

Date Ordered: {date_ordered}



------------------------------------------

Thank you for using our service!



"""



    messagebox.showinfo("Receipt", receipt)



def clear_fields():



    customer_entry.delete(0, tk.END)

    quantity_entry.delete(0, tk.END)

    price_entry.delete(0, tk.END)



    print_type_var.set("Select Print Type")

    paper_size_var.set("Select Paper Size")

    status_var.set("Select Status")





def submit_enter(event):

    submit_btn.config(bg="light green")



def submit_leave(event):

    submit_btn.config(bg="green")





def update_enter(event):

    update_btn.config(bg="orange")



def update_leave(event):

    update_btn.config(bg="yellow")





def delete_enter(event):

    delete_btn.config(bg="pink")



def delete_leave(event):

    delete_btn.config(bg="red")





def clear_enter(event):

    clear_btn.config(bg="black")



def clear_leave(event):

    clear_btn.config(bg="gray")





def receipt_enter(event):

    receipt_btn.config(bg="brown")



def receipt_leave(event):

    receipt_btn.config(bg="purple")



window = tk.Tk()

window.title("Printing Service Order System")

window.geometry("1200x700")

window.configure(bg="lightblue")

window.resizable(False, False)



style = ttk.Style()

style.theme_use("clam")



style.configure("Treeview",font=("Arial", 10),rowheight=30,background="white",foreground="black",fieldbackground="white")



style.configure("Treeview.Heading",font=("Arial", 11, "bold"),background="navy",foreground="white")



title = tk.Label(window,text="PRINTING SERVICE ORDER SYSTEM",font=("Arial", 24, "bold"),bg="lightblue",fg="darkblue")



title.pack(pady=15)





frame = tk.Frame(window,bg="white",bd=3,relief="ridge")

frame.pack(pady=10)



customer_entry = tk.Entry(frame,font=("Arial", 12),width=25,bd=2,relief="solid")



customer_entry.grid(row=0, column=0, padx=15, pady=10)



customer_label = tk.Label(frame,text="Customer Name",font=("Arial", 10, "bold"),bg="white",fg="darkblue")

customer_label.grid(row=1, column=0)



print_type_var = tk.StringVar()

print_type_var.set("Select Print Type")



printtype_menu = tk.OptionMenu(frame,print_type_var,"Black & White","Colored")

printtype_menu.config(width=20)

printtype_menu.grid(row=0, column=1, padx=15)



printtype_label = tk.Label(frame,text="Print Type",font=("Arial", 10, "bold"),bg="white",fg="darkblue")

printtype_label.grid(row=1, column=1)



paper_size_var = tk.StringVar()

paper_size_var.set("Select Paper Size")



papersize_menu = tk.OptionMenu(frame,paper_size_var,"Short","A4","Long")

papersize_menu.config(width=20)

papersize_menu.grid(row=0, column=2, padx=15)



papersize_label = tk.Label(frame,text="Paper Size",font=("Arial", 10, "bold"),bg="white",fg="darkblue")

papersize_label.grid(row=1, column=2)



quantity_entry = tk.Entry(frame,font=("Arial", 12),width=25,bd=2,relief="solid")

quantity_entry.grid(row=2, column=0, padx=15, pady=10)



quantity_label = tk.Label(frame,text="Quantity",font=("Arial", 10, "bold"),bg="white",fg="darkblue")

quantity_label.grid(row=3, column=0)



price_entry = tk.Entry(frame,font=("Arial", 12),width=25,bd=2,relief="solid")

price_entry.grid(row=2, column=1, padx=15)



price_label = tk.Label(frame,text="Price Per Page",font=("Arial", 10, "bold"),bg="white",fg="darkblue")

price_label.grid(row=3, column=1)



status_var = tk.StringVar()

status_var.set("Select Status")



status_menu = tk.OptionMenu(frame,status_var,"Pending","Done")

status_menu.config(width=20)

status_menu.grid(row=2, column=2, padx=15)



status_label = tk.Label(frame,text="Status",font=("Arial", 10, "bold"),bg="white",fg="darkblue")

status_label.grid(row=3, column=2)



button_frame = tk.Frame(window, bg="lightblue")

button_frame.pack(pady=15)



submit_btn = tk.Button(button_frame,text="ADD RECORD",bg="green",fg="white",font=("Arial", 11, "bold"),width=15,command=saving)

submit_btn.grid(row=0, column=0, padx=10)

submit_btn.bind("<Enter>", submit_enter)

submit_btn.bind("<Leave>", submit_leave)





update_btn = tk.Button(button_frame,text="UPDATE",bg="gold",fg="black",font=("Arial", 11, "bold"),width=15,command=update)

update_btn.grid(row=0, column=1, padx=10)

update_btn.bind("<Enter>", update_enter)

update_btn.bind("<Leave>", update_leave)





delete_btn = tk.Button(button_frame,text="DELETE",bg="red",fg="white",font=("Arial", 11, "bold"),width=15,command=delete)

delete_btn.grid(row=0, column=2, padx=10)

delete_btn.bind("<Enter>", delete_enter)

delete_btn.bind("<Leave>", delete_leave)





clear_btn = tk.Button(button_frame,text="CLEAR",bg="gray",fg="white",font=("Arial", 11, "bold"),width=15,command=clear_fields)

clear_btn.grid(row=0, column=3, padx=10)

clear_btn.bind("<Enter>", clear_enter)

clear_btn.bind("<Leave>", clear_leave)





receipt_btn = tk.Button(button_frame,text="PRINT RECEIPT",bg="purple",fg="white",font=("Arial", 11, "bold"),width=15,command=print_receipt)

receipt_btn.grid(row=0, column=4, padx=10)

receipt_btn.bind("<Enter>", receipt_enter)

receipt_btn.bind("<Leave>", receipt_leave)





search_frame = tk.Frame(window, bg="lightblue")

search_frame.pack(pady=10)



search_label = tk.Label(search_frame,text="Search:",font=("Arial", 11, "bold"),bg="lightblue",fg="darkblue")

search_label.grid(row=0, column=0, padx=10)



search_entry = tk.Entry(search_frame,font=("Arial", 11),width=30)

search_entry.grid(row=0, column=1)



search_btn = tk.Button(search_frame,text="SEARCH",bg="blue",fg="white",font=("Arial", 10, "bold"),command=search_record)

search_btn.grid(row=0, column=2, padx=10)



showall_btn = tk.Button(search_frame,text="SHOW ALL",bg="gray",fg="white",font=("Arial", 10, "bold"),command=display)

showall_btn.grid(row=0, column=3)



table_frame = tk.Frame(window)

table_frame.pack(pady=20)



scroll_y = tk.Scrollbar(table_frame, orient="vertical")

scroll_y.pack(side="right", fill="y")



table = ttk.Treeview(table_frame,columns=("Order ID","Customer Name","Print Type","Paper Size","Quantity","Price","Total","Status","Date Ordered"),show="headings",yscrollcommand=scroll_y.set,height=12)

scroll_y.config(command=table.yview)



headings = ("Order ID","Customer Name","Print Type","Paper Size","Quantity","Price","Total","Status","Date Ordered")



for head in headings:

    table.heading(head, text=head)

    table.column(head, width=130, anchor="center")



table.pack()



table.bind("<<TreeviewSelect>>", auto_populate)



display()

window.mainloop()